import openpyxl
import json
import subprocess
from pandas import DataFrame
import csv
excel_File_Path = "to_match_data.xlsx"

workbook_object = openpyxl.load_workbook(excel_File_Path)

sheet_object = workbook_object.worksheets[0]

sheet_object_3 = workbook_object.worksheets[1]
sheet_object_4 = workbook_object.worksheets[2]

rows = []
rows_1 = []
rows_2 = []
for row in sheet_object.iter_rows():
    lis = []
    

    for cell in row:
        lis.append(cell.value)
        
    rows.append(lis)

for i in sheet_object_3.iter_rows():
    lis = []
    

    for cell in i:
        lis.append(cell.value)
        
    rows_1.append(lis)

for j in sheet_object_4.iter_rows():
    lis = []
    

    for cell in j:
        lis.append(cell.value)
        
    rows_2.append(lis)

print(rows[1][0])  
print(rows_1[1][0]) 
list_1 = []
for i in range(len(rows)):
    list_1.append(str(rows[i][0]).lower())
    #host = rows[i][0]
    #ip = rows[i][16]
list_2 = []
for i in range(len(rows_1)):
    list_2.append(str(rows_1[i][0]).lower())

list_3 = []
for i in range(len(rows_2)):
    list_3.append(str(rows_2[i][0]).lower())

#print(list_2)






import requests, json

entities = []
API_URL = 'https://api.newrelic.com/graphql'
headers = { 'Api-Key': 'API-KEY' }

nextCursor = "null"
while True:
    #print('Next cursor:', nextCursor)
    query = """
    {
      actor {
        entitySearch(queryBuilder: {domain: INFRA, type: HOST}) {
          query
          results(cursor: NEXT_CURSOR) {
            nextCursor
            entities {
              name
            }
          }
        }
      }
    }
    """.replace('NEXT_CURSOR', nextCursor)

    response = requests.post(API_URL, headers = headers, json = { 'query': query })
    results = response.json()['data']['actor']['entitySearch']['results']
    entities += results['entities']
    if not results['nextCursor']: break
    nextCursor = f"\"{results['nextCursor']}\""
print(len(entities))
print(type(entities))

hosts = []

for i in entities:
    temp = i.get('name')
    hosts.append(str(temp.lower()))
#print(hosts)
#print(len(hosts))

matching_host_1 = []
matching_host_2 = []
matching_host_3 = []

not_matching_1 = []
not_matching_2 = []
not_matching_3 = []


for i in hosts:
    for j in list_1:
        if i in j:
            #print("true")
            #print(i)
            matching_host_1.append(j)
for i in hosts:
    for j in list_2:
        if i in j:
            matching_host_2.append(j)
for i in hosts:
    for j in list_3:
        if i in j:
            matching_host_3.append(j)

#print(len(list_1))
#print(len(matching_host_1))

for i in matching_host_1:
    for j in list_1:
        if i ==j:
            list_1.remove(j)
#print(len(list_1))
not_matching_1 = list_1

for i in matching_host_2:
    for j in list_2:
        if i ==j:
            list_2.remove(j)
not_matching_2 = list_2
for i in matching_host_3:
    for j in list_3:
        if i ==j:
            list_3.remove(j)
not_matching_3 = list_3
#print(matching_host_1)
#print(len(matching_host_1))
#print(len(matching_host_2))
#print("NEXTTTTT")
#print(matching_host_2)



#with open('entities.json', 'w') as entities_file:
 #   entities_file.write(json.dumps(entities))
df = DataFrame({'hosts': matching_host_1})
df_1 = DataFrame({'hosts': matching_host_2})
df_2 = DataFrame({'hosts': matching_host_3})
df.to_excel('cassamba.xlsx', sheet_name='sheet1', index=False)
df_1.to_excel('TherapyX.xlsx',sheet_name='sheet_1',index=False)
df_2.to_excel('TherapyO.xlsx',sheet_name='sheet1',index=False)

df_3= DataFrame({'not_maching': not_matching_1})
df_4=DataFrame({'not_maching': not_matching_2})
df_5=DataFrame({'not_maching': not_matching_3})
df_3.to_excel('cassamba_not_maching.xlsx', sheet_name='sheet1', index=False)
df_4.to_excel('TherapyX_not_maching.xlsx',sheet_name='sheet_1',index=False)
df_5.to_excel('TherapyO_not_maching.xlsx',sheet_name='sheet1',index=False)
